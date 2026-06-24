"""Excel round-trip of tournaments via TournamentResource (django-import-export).

Covers: export uses human-readable FK values, import creates rows and recomputes
the derived columns (buy_in_total / is_bounty / early_bird, plus verified_by_admin
which is recomputed but never exported), import updates by id without duplicating,
a clean round-trip reports no errors, and a series that belongs to a different room
is rejected as a row error.
"""

from datetime import timedelta
from decimal import Decimal
from io import BytesIO

import pytest
import tablib
from django.contrib.auth import get_user_model
from django.utils import timezone

from apps.rooms.models import Network, PokerRoom
from apps.tournaments.models import (
    BlindStructure,
    BlindStructureTemplate,
    BountyOption,
    BubbleOption,
    EarlyBirdType,
    GameType,
    Periodicity,
    ReEntryOption,
    Tournament,
    TournamentSeries,
)
from apps.tournaments.resources import COLUMN_LABELS as L
from apps.tournaments.resources import TournamentResource
from apps.tournaments.xlsx_export import LockedDropdownXLSX

User = get_user_model()


@pytest.fixture
def superuser():
    return User.objects.create_superuser(username="ie_admin", email="ie@example.com", password="x")


@pytest.fixture
def series():
    room = PokerRoom.objects.get(slug="pokerok")
    obj, _ = TournamentSeries.objects.get_or_create(
        room=room, slug="ie-series", defaults={"name": "IE Series"}
    )
    return obj


def _make_tournament(series, **overrides) -> Tournament:
    fields = dict(
        room=series.room,
        series=series,
        name="Daily NLHE",
        game_type=GameType.NLHE,
        buy_in_total=Decimal("55.00"),
        buy_in_without_rake=Decimal("50.00"),
        bounty_buyin=Decimal("0"),
        rake=Decimal("5.00"),
        guaranteed_dollars=10000,
        payout_percent=15,
        starting_stack=10000,
        starting_stack_bb=50,
        starting_time=timezone.now() + timedelta(hours=1),
        late_reg_at=timezone.now() + timedelta(hours=2),
        late_reg_level=12,
        blind_interval_minutes=10,
        break_minutes=5,
        players_per_table=9,
        players_at_final_table=9,
        min_players=2,
        max_players=1000,
        re_entry=ReEntryOption.objects.get(name="unlimited"),
        bubble=BubbleOption.objects.get(name="finalized_when_registration_closes"),
        early_bird=True,
        early_bird_type=EarlyBirdType.objects.get(name="compensated_at_bubble"),
        featured_final_table=False,
        periodicity=Periodicity.objects.get(name="one_off"),
        verified_by_admin=True,
    )
    fields.update(overrides)
    return Tournament.objects.create(**fields)


def _export_dataset(user, queryset) -> tablib.Dataset:
    return TournamentResource(user=user).export(queryset)


def _row_as_dict(dataset: tablib.Dataset, index: int = 0) -> dict:
    return dict(dataset.dict[index])


def _dataset_from_rows(rows: list[dict]) -> tablib.Dataset:
    ds = tablib.Dataset(headers=list(rows[0].keys()))
    for row in rows:
        ds.append(list(row.values()))
    return ds


@pytest.mark.django_db
def test_admin_import_export_pages_render(admin_client):
    # The ImportExportMixin adds these views; a 200 confirms the buttons are wired.
    assert admin_client.get("/admin/tournaments/tournament/import/").status_code == 200
    # Export skips the field-selection form: the button downloads the xlsx directly.
    resp = admin_client.get("/admin/tournaments/tournament/export/")
    assert resp.status_code == 200
    assert resp["Content-Type"] == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "attachment" in resp["Content-Disposition"]


@pytest.mark.django_db
def test_export_returns_rows(superuser, series):
    t = _make_tournament(series)
    ds = _export_dataset(superuser, Tournament.objects.filter(pk=t.pk))

    assert L["room"] in ds.headers and L["series"] in ds.headers
    row = _row_as_dict(ds)
    # Foreign keys are exported as the names an editor types, not ids.
    assert row[L["room"]] == series.room.name
    assert row[L["series"]] == series.name


@pytest.mark.django_db
def test_import_creates_tournament(superuser, series):
    template = _make_tournament(series)
    ds = _export_dataset(superuser, Tournament.objects.filter(pk=template.pk))
    row = _row_as_dict(ds)
    row[L["id"]] = ""  # blank id → create
    row[L["name"]] = "Imported NLHE"
    row[L["buy_in_total"]] = "999.00"  # deliberately wrong; must be recomputed

    result = TournamentResource(user=superuser).import_data(
        _dataset_from_rows([row]), dry_run=False
    )
    assert not result.has_errors(), result.row_errors()

    created = Tournament.objects.get(name="Imported NLHE")
    assert created.pk != template.pk
    # buy_in_total recomputed from parts (50 + 0 + 5), ignoring the bad cell.
    assert created.buy_in_total == Decimal("55.00")
    assert created.early_bird is True  # early_bird_type was set
    assert created.is_bounty is False  # no bounty_type
    assert created.verified_by_admin is True  # imported by a superuser


@pytest.mark.django_db
def test_import_derives_is_bounty(superuser, series):
    template = _make_tournament(series)
    ds = _export_dataset(superuser, Tournament.objects.filter(pk=template.pk))
    row = _row_as_dict(ds)
    row[L["id"]] = ""
    row[L["name"]] = "Bounty NLHE"
    row[L["bounty_buyin"]] = "10.00"
    row[L["bounty_type"]] = BountyOption.objects.first().name

    result = TournamentResource(user=superuser).import_data(
        _dataset_from_rows([row]), dry_run=False
    )
    assert not result.has_errors(), result.row_errors()

    created = Tournament.objects.get(name="Bounty NLHE")
    assert created.is_bounty is True


@pytest.mark.django_db
def test_import_updates_by_id(superuser, series):
    t = _make_tournament(series)
    ds = _export_dataset(superuser, Tournament.objects.filter(pk=t.pk))
    row = _row_as_dict(ds)
    row[L["name"]] = "Renamed"

    before = Tournament.objects.count()
    result = TournamentResource(user=superuser).import_data(
        _dataset_from_rows([row]), dry_run=False
    )
    assert not result.has_errors(), result.row_errors()

    assert Tournament.objects.count() == before  # updated, not duplicated
    t.refresh_from_db()
    assert t.name == "Renamed"


@pytest.mark.django_db
def test_round_trip(superuser, series):
    _make_tournament(series, name="A")
    _make_tournament(series, name="B")
    ds = _export_dataset(superuser, Tournament.objects.all())

    before = Tournament.objects.count()
    result = TournamentResource(user=superuser).import_data(ds, dry_run=False)

    assert not result.has_errors(), result.row_errors()
    assert Tournament.objects.count() == before  # unchanged rows, no new ones


@pytest.mark.django_db
def test_export_uses_admin_labels_locked_and_filterable(superuser, series):
    import openpyxl
    from openpyxl.utils import get_column_letter

    t = _make_tournament(series)
    dataset = _export_dataset(superuser, Tournament.objects.filter(pk=t.pk))
    content = LockedDropdownXLSX().export_data(dataset)

    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active
    headers = [c.value for c in ws[1]]

    # Columns are headed with the admin-form labels, not model field names.
    assert "ID" in headers and "Tournament series" in headers
    assert "id" not in headers and "series" not in headers
    room_col = headers.index(L["room"]) + 1

    # Sheet is protected, but AutoFilter is allowed (filtering works, sorting blocked).
    assert ws.protection.sheet is True
    assert ws.protection.autoFilter is False  # False == filtering allowed
    assert ws.auto_filter.ref and ws.auto_filter.ref.startswith("A1:")
    assert ws.freeze_panes == "A2"

    # Read-only columns are locked + grey-shaded; editable ones aren't.
    for field in ("id", "buy_in_total", "is_bounty", "early_bird"):
        col = headers.index(L[field]) + 1
        assert ws.cell(row=2, column=col).fill.fill_type == "solid", field
        assert ws.cell(row=2, column=col).protection.locked is True, field
    assert ws.cell(row=2, column=room_col).fill.fill_type in (None, "none")
    assert ws.cell(row=2, column=room_col).protection.locked is False
    assert ws.cell(row=1, column=room_col).protection.locked is True  # header locked

    # verified_by_admin is recomputed on import but never exported.
    assert L.get("verified_by_admin") is None  # not even mapped
    assert "Verified by admin" not in headers

    # The option columns carry a list validation pointing at the hidden sheet.
    assert "lists" in wb.sheetnames
    assert wb["lists"].sheet_state == "hidden"
    validated_ranges = " ".join(str(dv.sqref) for dv in ws.data_validations.dataValidation)
    assert get_column_letter(room_col) in validated_ranges

    # The instruction sheet is gone; the data sheet is active; headers have notes.
    assert "Инструкция" not in wb.sheetnames
    assert ws.title != "lists"
    assert ws.cell(row=1, column=room_col).comment is not None


@pytest.mark.django_db
def test_series_dropdown_cascades_from_room(superuser, series):
    import openpyxl
    from openpyxl.utils import get_column_letter

    t = _make_tournament(series)
    dataset = _export_dataset(superuser, Tournament.objects.filter(pk=t.pk))
    content = LockedDropdownXLSX().export_data(dataset)

    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    room_letter = get_column_letter(headers.index(L["room"]) + 1)
    series_letter = get_column_letter(headers.index(L["series"]) + 1)

    # The room has a named range over its series on the hidden sheet.
    assert series.room.name in wb.defined_names
    assert wb.defined_names[series.room.name].attr_text.startswith("lists!")

    # The series column validates against INDIRECT(<room cell>) — empty room → empty list.
    series_dvs = [
        dv
        for dv in ws.data_validations.dataValidation
        if any(rng.coord.startswith(series_letter) for rng in dv.cells.ranges)
    ]
    assert len(series_dvs) == 1
    assert series_dvs[0].formula1 == f"INDIRECT(${room_letter}2)"


def _give_blinds(tournament) -> None:
    BlindStructure.objects.create(
        tournament=tournament, level=1, small_blind=50, big_blind=100, ante=0
    )
    BlindStructure.objects.create(
        tournament=tournament, level=2, small_blind=100, big_blind=200, ante=25
    )


@pytest.mark.django_db
def test_export_includes_blind_structure_name(superuser, series):
    import openpyxl
    from openpyxl.utils import get_column_letter

    t = _make_tournament(series)
    _give_blinds(t)
    BlindStructureTemplate.create_from_tournament(t, name="Test Struct [abc123]")

    ds = _export_dataset(superuser, Tournament.objects.filter(pk=t.pk))
    assert _row_as_dict(ds)[L["blind_structure"]] == "Test Struct [abc123]"

    # The column also carries a dropdown of template names.
    content = LockedDropdownXLSX().export_data(ds)
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    bs_letter = get_column_letter(headers.index(L["blind_structure"]) + 1)
    validated = " ".join(str(dv.sqref) for dv in ws.data_validations.dataValidation)
    assert bs_letter in validated


@pytest.mark.django_db
def test_import_applies_blind_structure(superuser, series):
    source = _make_tournament(series)
    _give_blinds(source)
    BlindStructureTemplate.create_from_tournament(source, name="Import Struct [def456]")

    ds = _export_dataset(superuser, Tournament.objects.filter(pk=source.pk))
    row = _row_as_dict(ds)
    row[L["id"]] = ""
    row[L["name"]] = "Imported with blinds"
    row[L["blind_structure"]] = "Import Struct [def456]"

    result = TournamentResource(user=superuser).import_data(
        _dataset_from_rows([row]), dry_run=False
    )
    assert not result.has_errors(), result.row_errors()

    created = Tournament.objects.get(name="Imported with blinds")
    levels = list(
        created.blind_levels.order_by("level").values_list("small_blind", "big_blind", "ante")
    )
    assert levels == [(50, 100, 0), (100, 200, 25)]


@pytest.mark.django_db
def test_import_unknown_blind_structure_row_errors(superuser, series):
    template = _make_tournament(series)
    ds = _export_dataset(superuser, Tournament.objects.filter(pk=template.pk))
    row = _row_as_dict(ds)
    row[L["id"]] = ""
    row[L["name"]] = "Bad blinds"
    row[L["blind_structure"]] = "no-such-structure"

    result = TournamentResource(user=superuser).import_data(_dataset_from_rows([row]), dry_run=True)
    assert result.has_validation_errors()


@pytest.mark.django_db
def test_hardened_export_still_imports(superuser, series):
    """The locked/dropdown file round-trips through import unchanged."""
    import openpyxl

    _make_tournament(series, name="A")
    _make_tournament(series, name="B")
    dataset = _export_dataset(superuser, Tournament.objects.all())
    content = LockedDropdownXLSX().export_data(dataset)

    reparsed = LockedDropdownXLSX().create_dataset(content)
    before = Tournament.objects.count()
    result = TournamentResource(user=superuser).import_data(reparsed, dry_run=False)

    assert not result.has_errors(), result.row_errors()
    assert Tournament.objects.count() == before  # updated in place, no duplicates
    # The hidden helper sheet must not leak into the parsed data.
    assert openpyxl.load_workbook(BytesIO(content)).active.title != "lists"


@pytest.mark.django_db
def test_import_bad_series_row_errors(superuser, series):
    # A series that belongs to a *different* room must not resolve for pokerok.
    other_net, _ = Network.objects.get_or_create(slug="othernet", defaults={"name": "OtherNet"})
    other_room = PokerRoom.objects.create(name="OtherRoom", slug="otherroom", network=other_net)
    TournamentSeries.objects.create(room=other_room, slug="foreign", name="Foreign Series")

    template = _make_tournament(series)
    ds = _export_dataset(superuser, Tournament.objects.filter(pk=template.pk))
    row = _row_as_dict(ds)
    row[L["id"]] = ""
    row[L["name"]] = "Cross-room"
    row[L["series"]] = "Foreign Series"  # exists, but not in pokerok

    result = TournamentResource(user=superuser).import_data(_dataset_from_rows([row]), dry_run=True)
    # A widget ValueError surfaces as a per-row validation error, which blocks
    # the import (the row is reported back to the editor rather than created).
    assert result.has_validation_errors()
    assert result.invalid_rows[0].error_count == 1


@pytest.mark.django_db
def test_import_rejects_file_with_missing_columns(superuser):
    # A file that isn't a tournament export (wrong/missing headers) is rejected
    # up front with a clear base error rather than failing per row.
    bad = tablib.Dataset(headers=["id", "room", "name"])
    bad.append(["", "PokerStars", "x"])

    result = TournamentResource(user=superuser).import_data(bad, dry_run=True, raise_errors=False)
    assert result.base_errors
    assert "does not match the tournament export format" in str(result.base_errors[0].error)
