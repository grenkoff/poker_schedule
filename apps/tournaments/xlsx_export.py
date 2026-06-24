"""Custom XLSX export format that hardens the hand-editable tournament file.

`django-import-export` produces a plain xlsx via tablib; tablib can't lock cells
or attach dropdowns. So this format runs the tablib output back through openpyxl
to add two editor affordances, then hands the bytes onward unchanged:

* The ``id`` column, the columns recomputed on import (``buy_in_total``,
  ``is_bounty``, ``early_bird``), and the header row are locked (sheet protection
  on, every other cell — plus a buffer of empty rows for new tournaments — left
  editable). ``id`` is the import match key (see
  ``TournamentResource.import_id_fields``) so hand-editing it silently
  overwrites the wrong row; the computed columns are overwritten on import so
  editing them has no effect; renaming a header breaks the column→field mapping.
* Columns backed by a fixed option set get a data-validation dropdown listing the
  exact strings import accepts — FK columns export the option ``name`` slug,
  ``game_type`` exports the choice code. The option values live on a hidden
  helper sheet so we're not bound by the 255-char inline-list limit.
* ``series`` is a *cascading* dropdown scoped to the row's ``room``: each room
  gets a named range over its series on the helper sheet and the series cell
  validates against ``INDIRECT(<room cell>)``, so it stays empty until a room is
  picked and then offers only that room's series.
* Read-only columns are shaded grey, each header carries a hover note, and a
  visible "Инструкция" sheet spells out the legend for whoever fills the file.

Import is untouched: the parsing path (`create_dataset`) ignores protection,
validation, fills and extra sheets, so a file produced here round-trips through
import unchanged (it reads only the active data sheet's header + rows).
"""

from __future__ import annotations

import re
from io import BytesIO

from import_export.formats import base_formats

# A workbook-defined name must start with a letter/underscore, hold only
# letters/digits/_/. and not look like a cell reference (e.g. "AB12"). Room
# names are used verbatim as range names for the cascading series dropdown, so
# any room whose name fails this is skipped (its series simply won't cascade).
_VALID_NAME_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_.]*$")
_CELL_REF_RE = re.compile(r"^[A-Za-z]{1,3}[0-9]+$")

# Columns whose valid values are a closed set. `series` is deliberately absent:
# its options depend on the row's `room`, so a flat dropdown would offer
# wrong-room series. Booleans are absent too — they export as "1"/"0".
_DROPDOWN_COLUMNS = (
    "room",
    "game_type",
    "re_entry",
    "bubble",
    "periodicity",
    "bounty_type",
    "early_bird_type",
    "deal_making",
    "blind_structure",
)

# Columns recomputed on import (see TournamentResource.before_save_instance), so
# hand-editing them has no effect — lock them alongside `id` to make that clear.
_COMPUTED_COLUMNS = (
    "buy_in_total",
    "is_bounty",
    "early_bird",
)

# Extra empty rows below the data kept editable (and dropdown-equipped) so an
# editor can append new tournaments under a protected sheet.
_EXTRA_ROWS = 200

# Light grey shading for the read-only (locked) columns, so the lock is visible
# at a glance and matches the legend on the "Инструкция" sheet.
_LOCKED_FILL = "D9D9D9"


def _dropdown_options() -> dict[str, list[str]]:
    """Allowed cell strings per dropdown column, in display order.

    Imported lazily so this module stays import-safe before the app registry is
    ready (it's referenced from admin at class-definition time).
    """
    from apps.rooms.models import PokerRoom

    from .models import (
        BlindStructureTemplate,
        BountyOption,
        BubbleOption,
        DealMakingOption,
        EarlyBirdType,
        GameType,
        Periodicity,
        ReEntryOption,
    )

    def _names(model) -> list[str]:
        return list(model.objects.order_by("sort_order", "label").values_list("name", flat=True))

    return {
        "room": list(PokerRoom.objects.order_by("name").values_list("name", flat=True)),
        "game_type": [choice.value for choice in GameType],
        "re_entry": _names(ReEntryOption),
        "bubble": _names(BubbleOption),
        "periodicity": _names(Periodicity),
        "bounty_type": _names(BountyOption),
        "early_bird_type": _names(EarlyBirdType),
        "deal_making": _names(DealMakingOption),
        "blind_structure": list(
            BlindStructureTemplate.objects.order_by("name").values_list("name", flat=True)
        ),
    }


def _series_by_room() -> dict[str, list[str]]:
    """Series names grouped by room name, in display order.

    Drives the cascading ``series`` dropdown; every series (including the legacy
    "Default") is kept so existing exported values stay valid against the list.
    """
    from collections import defaultdict

    from .models import TournamentSeries

    grouped: dict[str, list[str]] = defaultdict(list)
    qs = TournamentSeries.objects.select_related("room").order_by(
        "room__name", "sort_order", "name"
    )
    for series in qs:
        grouped[series.room.name].append(series.name)
    return grouped


def _valid_range_name(name: str) -> str | None:
    """Return ``name`` if it's usable verbatim as an Excel defined name, else None."""
    if len(name) > 255 or not _VALID_NAME_RE.fullmatch(name) or _CELL_REF_RE.fullmatch(name):
        return None
    if name.upper() in {"R", "C"}:  # reserved single-letter names
        return None
    return name


def _harden_workbook(content: bytes) -> bytes:
    import openpyxl
    from openpyxl.comments import Comment
    from openpyxl.styles import PatternFill, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    col_of = {name: idx + 1 for idx, name in enumerate(headers) if name}
    last_row = ws.max_row + _EXTRA_ROWS

    # --- dropdowns on closed-set columns ---------------------------------
    options = _dropdown_options()
    lists = wb.create_sheet("lists")
    lists.sheet_state = "hidden"
    next_list_col = 1  # next free column on the hidden helper sheet
    for column_name in _DROPDOWN_COLUMNS:
        values = options.get(column_name)
        target_col = col_of.get(column_name)
        if not values or not target_col:
            continue
        letter = get_column_letter(next_list_col)
        lists.cell(row=1, column=next_list_col, value=column_name)
        for offset, value in enumerate(values, start=2):
            lists.cell(row=offset, column=next_list_col, value=value)
        ref = f"lists!${letter}$2:${letter}${len(values) + 1}"
        dv = DataValidation(type="list", formula1=ref, allow_blank=True)
        ws.add_data_validation(dv)
        target_letter = get_column_letter(target_col)
        dv.add(f"{target_letter}2:{target_letter}{last_row}")
        next_list_col += 1

    # --- cascading series dropdown, scoped to the row's room -------------
    series_col = col_of.get("series")
    room_col = col_of.get("room")
    if series_col and room_col:
        from openpyxl.workbook.defined_name import DefinedName

        for room_name, series_names in _series_by_room().items():
            safe = _valid_range_name(room_name)
            if not series_names or safe is None:
                continue
            letter = get_column_letter(next_list_col)
            lists.cell(row=1, column=next_list_col, value=room_name)
            for offset, value in enumerate(series_names, start=2):
                lists.cell(row=offset, column=next_list_col, value=value)
            ref = f"lists!${letter}$2:${letter}${len(series_names) + 1}"
            wb.defined_names.add(DefinedName(safe, attr_text=ref))
            next_list_col += 1

        # INDIRECT resolves the room cell value to its named range; an empty
        # room → empty list, so series stays inactive until a room is chosen.
        room_letter = get_column_letter(room_col)
        series_letter = get_column_letter(series_col)
        series_dv = DataValidation(
            type="list", formula1=f"INDIRECT(${room_letter}2)", allow_blank=True
        )
        ws.add_data_validation(series_dv)
        series_dv.add(f"{series_letter}2:{series_letter}{last_row}")

    # --- lock id + computed columns + header, grey-shade read-only cols --
    locked_cols = {col_of[name] for name in ("id", *_COMPUTED_COLUMNS) if name in col_of}
    grey = PatternFill(start_color=_LOCKED_FILL, end_color=_LOCKED_FILL, fill_type="solid")
    unlocked = Protection(locked=False)
    locked = Protection(locked=True)
    for row in range(1, last_row + 1):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            in_locked_col = col in locked_cols
            cell.protection = locked if (row == 1 or in_locked_col) else unlocked
            if in_locked_col:
                cell.fill = grey
    ws.protection.sheet = True
    ws.freeze_panes = "A2"  # keep headers visible while scrolling

    # --- per-header hover notes ------------------------------------------
    notes = {name: "Выберите значение из выпадающего списка." for name in _DROPDOWN_COLUMNS}
    notes.update(
        {
            name: "Только чтение: пересчитывается автоматически при импорте."
            for name in _COMPUTED_COLUMNS
        }
    )
    notes["id"] = (
        "Только чтение. Оставьте пустым для нового турнира; "
        "заполненный id обновляет существующий турнир."
    )
    notes["series"] = (
        "Сначала выберите room — затем здесь появится список серий этой комнаты. "
        "Пока room пустой, список недоступен."
    )
    notes["blind_structure"] = (
        "Выберите название структуры блайндов из списка. При импорте её уровни "
        "подставятся в турнир автоматически."
    )
    notes["starting_time"] = "Формат: ГГГГ-ММ-ДД ЧЧ:ММ, например 2026-06-22 19:30."
    notes["late_reg_at"] = notes["starting_time"]
    for name, text in notes.items():
        header_col = col_of.get(name)
        if header_col:
            ws.cell(row=1, column=header_col).comment = Comment(text, "schedule")

    _add_instruction_sheet(wb)
    wb.active = wb.index(ws)  # data sheet stays active so import reads it

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def _add_instruction_sheet(wb) -> None:
    """A visible "Инструкция" sheet with the legend for whoever fills the file."""
    from openpyxl.styles import Alignment, Font

    # (text, is_heading) — blank strings are spacer rows.
    lines: list[tuple[str, bool]] = [
        ("Как заполнять таблицу турниров", True),
        ("", False),
        (
            "Каждая строка — один турнир. Чтобы добавить турниры, дописывайте строки "
            "вниз (ниже уже подготовлены пустые строки со списками).",
            False,
        ),
        ("", False),
        ("Серый фон = только для чтения.", True),
        (
            "Серые колонки (id, buy_in_total, is_bounty, early_bird) заблокированы "
            "и заполняются автоматически — менять их вручную бесполезно.",
            False,
        ),
        ("", False),
        ("Колонка id", True),
        (
            "Оставьте пустой — будет создан новый турнир. Если id заполнен, обновится "
            "существующий турнир с этим id. Не вписывайте id вручную.",
            False,
        ),
        ("", False),
        ("Колонки с выпадающим списком", True),
        (
            "room, game_type, re_entry, bubble, periodicity, bounty_type, early_bird_type, "
            "deal_making, blind_structure — выбирайте значение из списка, не вводите вручную.",
            False,
        ),
        ("", False),
        ("blind_structure", True),
        (
            "Выберите название структуры блайндов. При импорте уровни этой структуры "
            "автоматически подставятся в турнир.",
            False,
        ),
        ("", False),
        ("series", True),
        (
            "Зависит от room: сначала выберите комнату — затем в series появится "
            "список серий этой комнаты. Пока room пустой, выбор серии недоступен.",
            False,
        ),
        ("", False),
        ("Дата и время (starting_time, late_reg_at)", True),
        ("Формат ГГГГ-ММ-ДД ЧЧ:ММ, например 2026-06-22 19:30.", False),
        ("", False),
        ("Деньги (buy_in_without_rake, bounty_buyin, rake)", True),
        ("Указывайте в долларах. buy_in_total считается автоматически как их сумма.", False),
        ("", False),
        (
            "Остальные колонки заполняйте вручную. Не переименовывайте заголовки — "
            "по ним идёт импорт.",
            False,
        ),
    ]
    sheet = wb.create_sheet("Инструкция", 1)  # second tab, right after the data
    for idx, (text, is_heading) in enumerate(lines, start=1):
        cell = sheet.cell(row=idx, column=1, value=text)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if is_heading:
            cell.font = Font(bold=True)
    sheet.column_dimensions["A"].width = 100


class LockedDropdownXLSX(base_formats.XLSX):
    """XLSX export with a locked ``id``/header and option dropdowns.

    Inherits the import path unchanged; only the export bytes are reworked.
    """

    def export_data(self, dataset, **kwargs):
        content = super().export_data(dataset, **kwargs)
        return _harden_workbook(content)
